import shutil
import zipfile
import os
import xml.etree.ElementTree as ET
import openpyxl
import pandas as pd

image_list = []
image_dict = {}

def read_excel_data(filename_path, sheet_name):
    """读取Excel数据并提取图片ID"""
    workbook = openpyxl.load_workbook(filename_path, data_only=False)
    sheet = workbook[sheet_name] if isinstance(sheet_name, str) else workbook.worksheets[sheet_name]

    data = []
    for row in sheet.iter_rows(min_row=1, values_only=False):
        row_data = []
        for cell in row:
            if cell.value and isinstance(cell.value, str) and '=_xlfn.DISPIMG(' in cell.value:
                formula = cell.value
                start = formula.find('"') + 1
                end = formula.find('"', start)
                image_id = formula[start:end]
                row_data.append(image_id)
                image_list.append(image_id)
            else:
                row_data.append(cell.value)
        data.append(row_data)
    workbook.close()
    return data

def get_xml_id_image_map(xlsx_file_path):
    """获取XML中图片ID和路径的映射"""
    with zipfile.ZipFile(xlsx_file_path, 'r') as zfile:
        with zfile.open('xl/cellimages.xml') as file:
            xml_content = file.read()
        with zfile.open('xl/_rels/cellimages.xml.rels') as file:
            relxml_content = file.read()

    root = ET.fromstring(xml_content)
    namespaces = {
        'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
        'a': 'http://schemas.openxmlformats.org/drawingml/2006/main'
    }

    name_to_embed_map = {}
    for pic in root.findall('.//xdr:pic', namespaces=namespaces):
        name = pic.find('.//xdr:cNvPr', namespaces=namespaces).attrib['name']
        embed = pic.find('.//a:blip', namespaces=namespaces).attrib[
            '{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed']
        name_to_embed_map[name] = embed

    root1 = ET.fromstring(relxml_content)
    namespaces = {'r': 'http://schemas.openxmlformats.org/package/2006/relationships'}

    id_target_map = {child.attrib['Id']: child.attrib.get('Target', 'No Target Found')
                     for child in root1.findall('.//r:Relationship', namespaces=namespaces)}

    name_to_target_map = {name: id_target_map[embed]
                         for name, embed in name_to_embed_map.items()
                         if embed in id_target_map}
    return name_to_target_map

def create_image_dict(xlsx_file_path, sheet_name, output_folder='images'):
    """创建图片ID和文件路径的映射字典"""
    read_excel_data(xlsx_file_path, sheet_name)
    name_to_target_map = get_xml_id_image_map(xlsx_file_path)

    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    with zipfile.ZipFile(xlsx_file_path, 'r') as zfile:
        for image_id in image_list:
            if image_id in name_to_target_map:
                image_path = name_to_target_map[image_id]
                actual_image_path = f'xl/{image_path}'
                if actual_image_path in zfile.namelist():
                    new_file_path = os.path.join(output_folder, f"{image_id}.png")
                    with zfile.open(actual_image_path) as image_file:
                        image_content = image_file.read()
                        with open(new_file_path, 'wb') as new_file:
                            new_file.write(image_content)
                    image_dict[image_id] = f"{output_folder}/{image_id}.png"

def clean_images_folder(folder_path):
    """清空图片文件夹"""
    if os.path.exists(folder_path):
        shutil.rmtree(folder_path)
    os.makedirs(folder_path)

def save_to_md(file_path, content):
    """保存内容到Markdown文件"""
    with open(file_path, 'w', encoding='utf-8') as f:
        f.write(content)

def excel_to_md(excel_file, sheet_name=0, output_folder='images'):
    """将Excel转换为Markdown"""
    global image_list, image_dict
    image_list.clear()
    image_dict.clear()

    create_image_dict(excel_file, sheet_name, output_folder)

    wb = openpyxl.load_workbook(excel_file, data_only=True)
    sheet = wb[sheet_name] if isinstance(sheet_name, str) else wb.worksheets[sheet_name]
    merged_ranges = sheet.merged_cells.ranges

    df = pd.read_excel(excel_file, sheet_name=sheet_name, engine='openpyxl')

    column_mapping = {
        '业务场景': '业务场景',
        '一级节点': '一级节点',
        '二级节点': '二级节点',
        '三级节点': '三级节点',
        '末级工作节点': '末级工作节点',
        '产出物名称': '产出物名称',
        '输入物样式': '输入物样式',
        '产出物样式': '产出物样式',
        '处理业务规则': '处理业务规则'
    }

    def get_merged_value(row_idx, col_name):
        try:
            # Convert to 1-based indices
            col_idx = df.columns.get_loc(col_name) + 1
            row_idx = row_idx + 2  # Excel row is 1-based and has header

            for merged_range in merged_ranges:
                if (merged_range.min_row <= row_idx <= merged_range.max_row and
                    merged_range.min_col <= col_idx <= merged_range.max_col):
                    cell_value = sheet.cell(merged_range.min_row, merged_range.min_col).value
                    return str(cell_value) if cell_value is not None else ''
            return None
        except Exception as e:
            print(f"Error in get_merged_value: row={row_idx}, col={col_name}, error={e}")
            return None

    def get_last_non_empty_node(current_nodes):
        for level in ['末级工作节点', '三级节点', '二级节点', '一级节点', '业务场景']:
            if current_nodes.get(level):
                return level, current_nodes[level]
        return None, None

    def get_second_last_non_empty_node(current_nodes):
        found_one = False
        for level in ['末级工作节点', '三级节点', '二级节点', '一级节点', '业务场景']:
            if current_nodes.get(level):
                if found_one:
                    return level, current_nodes[level]
                found_one = True
        return None, None

    def should_hide_heading(df, index, current_nodes, column_mapping):
        last_node_level, last_node_value = get_last_non_empty_node(current_nodes)
        second_last_level, second_last_value = get_second_last_non_empty_node(current_nodes)

        if not last_node_value or not second_last_value:
            return False, None

        # 如果当前值和上一级值相同
        if last_node_value == second_last_value:
            # 检查下一行的值
            next_index = index + 1
            if next_index < len(df):
                # 获取下一行的值（考虑合并单元格）
                next_merged_value = get_merged_value(next_index, column_mapping[second_last_level])
                if next_merged_value is None:
                    next_value = df.loc[next_index, column_mapping[second_last_level]]
                    next_value = str(next_value) if not pd.isna(next_value) else ''
                else:
                    next_value = next_merged_value

                # 如果倒数第二级节点的下一行的值与其不同，则隐藏当前节点
                if next_value != second_last_value:
                    return True, last_node_level
            else:
                return True, last_node_level
        else:
            return False, None

        return False, None

    markdown_content = ""
    prev_content = {key: '' for key in ['业务场景', '一级节点', '二级节点', '三级节点', '末级工作节点']}

    def process_nodes(df, index, current_nodes, column_mapping, prev_content):
        """递归处理需要隐藏的节点，返回最后一个被隐藏的节点信息"""
        last_hidden = None
        should_hide, node_to_hide = should_hide_heading(df, index, current_nodes, column_mapping)

        if should_hide and node_to_hide:
            current_nodes[node_to_hide] = ''
            prev_content[node_to_hide] = ''
            last_node_level, last_node_value = get_last_non_empty_node(current_nodes)
            last_hidden = (node_to_hide, df.loc[index, column_mapping[last_node_level]])

            # 继续递归检查其他节点
            deeper_hidden = process_nodes(df, index, current_nodes, column_mapping, prev_content)
            if deeper_hidden:
                last_hidden = deeper_hidden

        return last_hidden

    # 在主循环中修改为：
    for index, row in df.iterrows():
        current_nodes = {}

        # 获取当前行的所有节点值
        for node_type in ['业务场景', '一级节点', '二级节点', '三级节点', '末级工作节点']:
            merged_value = get_merged_value(index, column_mapping[node_type])
            if merged_value is not None:
                current_nodes[node_type] = merged_value
            else:
                value = row[column_mapping[node_type]]
                current_nodes[node_type] = str(value) if not pd.isna(value) else ''

        # if sheet_name=='环保工程':
        #     print(f"当前行的业务场景: {current_nodes['业务场景']}")
        # 处理需要隐藏的节点，并获取最后一个被隐藏的节点信息
        last_hidden = process_nodes(df, index, current_nodes, column_mapping, prev_content)

        # # 如果有节点被隐藏，恢复最后一个被隐藏的节点，保证正确的层级结构显示
        # if last_hidden:
        #     node_type, output_value = last_hidden
        #     current_nodes[node_type] = output_value
        #     prev_content[node_type] = ''  # 重置前一个内容，确保节点会被输出

        last_non_empty_value = ''
        for level in ['业务场景', '一级节点', '二级节点', '三级节点']:
            if current_nodes[level]:
                last_non_empty_value = current_nodes[level]
        if current_nodes['末级工作节点'] == last_non_empty_value:
            current_nodes['末级工作节点'] = ''
            prev_content['末级工作节点'] = ''

        last_level = None
        last_content = None
        for level, content in current_nodes.items():
            if content:
                last_level = level
                last_content = content

        # 处理节点输出
        for level, content in current_nodes.items():
            if content and content != prev_content[level]:
                if level == '业务场景':
                    markdown_content += f"# {content}\n"
                elif level == '一级节点':
                    markdown_content += f"## {content}\n"
                elif level == '二级节点':
                    markdown_content += f"### {content}\n"
                elif level == '三级节点':
                    markdown_content += f"#### {content}\n"
                elif level == '末级工作节点':
                    markdown_content += f"##### {content}\n"
                prev_content[level] = content

        # 处理产出物
        output_name = str(row[column_mapping['产出物名称']]) if not pd.isna(row[column_mapping['产出物名称']]) else ''
        if output_name == "设计井数情况信息表" and sheet_name == '节能评价':
            print(f"当前行的产出物名称为: {output_name}")
        if output_name == last_content:
            next_index = index + 1
            if next_index < len(df):
                next_merged_value = get_merged_value(next_index, column_mapping[last_level])
                if next_merged_value is None:
                    next_value = df.loc[next_index, column_mapping[last_level]]
                    next_value = str(next_value) if not pd.isna(next_value) else ''
                else:
                    next_value = next_merged_value

                if next_value == last_content:
                    markdown_content += f"- {output_name}\n"
        else:
            markdown_content += f"- {output_name}\n"


        input_style = str(row[column_mapping['输入物样式']]) if not pd.isna(row[column_mapping['输入物样式']]) else ''
        if input_style:
            markdown_content += "  - 输入物\n"
            if '=DISPIMG' in input_style:
                img_id = input_style[input_style.find('"') + 1:input_style.rfind('"')]
                if img_id in image_dict:
                    markdown_content += f"    - ![](images/{img_id}.png)\n"
                else:
                    markdown_content += f"    - {input_style}\n"
            else:
                markdown_content += f"    - {input_style}\n"

        output_style = str(row[column_mapping['产出物样式']]) if not pd.isna(row[column_mapping['产出物样式']]) else ''
        if output_style:
            markdown_content += "  - 产出物样式\n"
            if '=DISPIMG' in output_style:
                img_id = output_style[output_style.find('"') + 1:output_style.rfind('"')]
                if img_id in image_dict:
                    markdown_content += f"    - ![](images/{img_id}.png)\n"
                else:
                    markdown_content += f"    - {output_style}\n"
            else:
                markdown_content += f"    - {output_style}\n"

        processing_rules = str(row[column_mapping['处理业务规则']]) if not pd.isna(row[column_mapping['处理业务规则']]) else ''
        if processing_rules and processing_rules != 'nan':
            markdown_content += f"  - 处理业务规则\n    - {processing_rules}\n"
        last_hidden = None

    wb.close()
    return markdown_content

if __name__ == '__main__':
    excel_file = '安全环保IPOM表-20250224-生董董.xlsx'
    base_output_folder = 'output'
    images_folder = os.path.join(base_output_folder, 'images')

    try:
        if not os.path.exists(base_output_folder):
            os.makedirs(base_output_folder)

        clean_images_folder(images_folder)

        wb = openpyxl.load_workbook(excel_file, data_only=False)
        sheet_names = wb.sheetnames
        wb.close()

        for sheet_name in sheet_names:
            markdown_file = os.path.join(base_output_folder, f'{sheet_name}.md')
            print(f"正在处理工作表: {sheet_name}")
            md_content = excel_to_md(excel_file, sheet_name=sheet_name, output_folder=images_folder)
            save_to_md(markdown_file, md_content)
            print(f"已保存工作表 {sheet_name} 的内容到 {markdown_file}")

        print("所有工作表处理完成")
        print(f"输出目录: {os.path.abspath(base_output_folder)}")

    except Exception as e:
        print(f"发生错误：{e}")
        print("请检查Excel文件中的列名是否正确")