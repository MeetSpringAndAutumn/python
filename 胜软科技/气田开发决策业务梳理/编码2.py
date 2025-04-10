import openpyxl

def generate_encoding(file_path, encoding_column='AC'):
    """
    生成业务编码并写入 Excel 文件的指定列。

    :param file_path: Excel 文件路径
    :param encoding_column: 编码写入的列名（默认为 'AC'）
    """
    try:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        merged_ranges = sheet.merged_cells.ranges

        # 定义列名映射
        column_names = {
            '一级业务': None,
            '二级业务': None,
            '三级业务': None,
            '四级业务': None
        }

        # 在第一行查找列的索引
        for cell in sheet[1]:
            if cell.value in column_names:
                column_names[cell.value] = cell.column_letter

        # 确保所有必需的列都存在
        if None in column_names.values():
            missing_columns = [name for name, col in column_names.items() if col is None]
            raise ValueError(f"未找到必需的列: {', '.join(missing_columns)}")

        # 记录各级业务的编号
        level1_count = {}
        level2_count = {}
        level3_count = {}
        level4_count = {}

        # 遍历每一行，生成编码
        for row_idx in range(2, sheet.max_row + 1):
            level1 = level2 = level3 = level4 = None

            # 获取各级业务内容
            for name, col in column_names.items():
                cell = sheet[f'{col}{row_idx}']

                # 检查是否为合并单元格
                is_merged = False
                for merged_range in merged_ranges:
                    if cell.coordinate in merged_range:
                        start_cell = sheet.cell(merged_range.min_row, merged_range.min_col)
                        if name == '一级业务':
                            level1 = start_cell.value
                        elif name == '二级业务':
                            level2 = start_cell.value
                        elif name == '三级业务':
                            level3 = start_cell.value
                        elif name == '四级业务':
                            level4 = start_cell.value
                        is_merged = True
                        break

                # 如果不是合并单元格
                if not is_merged:
                    if name == '一级业务':
                        level1 = cell.value
                    elif name == '二级业务':
                        level2 = cell.value
                    elif name == '三级业务':
                        level3 = cell.value
                    elif name == '四级业务':
                        level4 = cell.value

            # 初始化各级业务的编号
            level1_code = level2_code = level3_code = level4_code = '0'

            # 一级业务编码
            if level1:
                if level1 not in level1_count:
                    level1_count[level1] = len(level1_count) + 1
                    level2_count = {}
                level1_code = str(level1_count[level1])

            # 二级业务编码
            if level2:
                key = (level1, level2)
                if key not in level2_count:
                    level2_count[key] = len(level2_count) + 1
                    level3_count = {}
                level2_code = str(level2_count[key])

            # 三级业务编码
            if level3:
                key = (level1, level2, level3)
                if key not in level3_count:
                    level3_count[key] = len(level3_count) + 1
                    level4_count = {}
                level3_code = str(level3_count[key])

            # 四级业务编码
            if level4:
                key = (level1, level2, level3, level4)
                if key not in level4_count:
                    level4_count[key] = len(level4_count) + 1
                level4_code = str(level4_count[key])

            # 生成最终编码
            encoding = f"{level1_code}{level2_code}{level3_code}{level4_code}"
            sheet[f'{encoding_column}{row_idx}'] = encoding

        wb.save(file_path)
        print("编码生成完成！")

    except PermissionError as e:
        print(f"权限错误：无法保存文件 '{file_path}'。请确保文件未被占用且具有写入权限。")
        print(e)
    except Exception as e:
        print(f"发生意外错误：{e}")

def generate_second_encoding(file_path, type_column=None, first_code_column='AC', second_code_column='AD'):
    try:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active

        # Find the "产出物类型" column
        type_col_letter = None
        for cell in sheet[1]:
            if cell.value == "产出物类型":
                type_col_letter = cell.column_letter
                break

        if not type_col_letter:
            raise ValueError("未找到'产出物类型'列")

        # Dictionary for first digit mapping
        type_mapping = {
            "图件": "T",
            "文档": "M",
            "结构化数据": "B",
            "公式": "S"
        }

        # Dictionary to track counters for each AC code + type combination
        counter_dict = {}

        # Generate codes for each row
        for row_idx in range(2, sheet.max_row + 1):
            ac_code = sheet[f'{first_code_column}{row_idx}'].value
            output_type = sheet[f'{type_col_letter}{row_idx}'].value

            if ac_code and output_type:
                # Get first digit based on type
                first_digit = type_mapping.get(output_type, "X")

                # Create key for counter
                key = f"{ac_code}_{first_digit}"

                # Initialize counter if not exists
                if key not in counter_dict:
                    counter_dict[key] = 0

                # Increment counter
                counter_dict[key] += 1

                # Generate new code
                new_code = f"{first_digit}{counter_dict[key]}"

                # Write new code to AD column
                sheet[f'{second_code_column}{row_idx}'] = new_code

        wb.save(file_path)
        print("第二组编码生成完成！")

    except PermissionError as e:
        print(f"权限错误：无法保存文件 '{file_path}'。请确保文件未被占用且具有写入权限。")
        print(e)
    except Exception as e:
        print(f"发生意外错误：{e}")

def main():
    file_path = '项目组_4  开发及其输入和输出--地面-2025.03.18-2（3.0环境导出）--Z 11.13 - 地面总体布局设计 ---最终.xlsx'
    generate_encoding(file_path)
    generate_second_encoding(file_path)

if __name__ == "__main__":
    main()