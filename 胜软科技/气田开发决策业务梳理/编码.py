import openpyxl

def generate_encoding(file_path, encoding_column='AC'):
    """
    生成业务编码并写入 Excel 文件的指定列。

    :param file_path: Excel 文件路径
    :param encoding_column: 编码写入的列名（默认为 'AC'）
    """
    try:
        # 打开 Excel 文件
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active

        # 获取合并单元格的范围
        merged_ranges = sheet.merged_cells.ranges

        # 记录各级业务的编号
        level1_count = {}
        level2_count = {}
        level3_count = {}
        level4_count = {}

        # 遍历每一行，生成编码
        for row in sheet.iter_rows(min_row=2):  # 从第二行开始，跳过标题行
            # 初始化各级业务的值
            level1 = None
            level2 = None
            level3 = None
            level4 = None

            # 获取各级业务内容（Cell 对象）
            for i, cell in enumerate(row[:4]):  # 只处理前四列
                # 检查当前单元格是否属于合并单元格
                is_merged = False
                for merged_range in merged_ranges:
                    if cell.coordinate in merged_range:
                        # 获取合并单元格的起始单元格
                        start_cell = sheet.cell(merged_range.min_row, merged_range.min_col)
                        if i == 0:
                            level1 = start_cell.value
                        elif i == 1:
                            level2 = start_cell.value
                        elif i == 2:
                            level3 = start_cell.value
                        elif i == 3:
                            level4 = start_cell.value
                        is_merged = True
                        break

                # 如果不是合并单元格，则直接使用当前单元格的值
                if not is_merged:
                    if i == 0:
                        level1 = cell.value
                    elif i == 1:
                        level2 = cell.value
                    elif i == 2:
                        level3 = cell.value
                    elif i == 3:
                        level4 = cell.value

            # 初始化各级业务的编号
            level1_code = '0'
            level2_code = '0'
            level3_code = '0'
            level4_code = '0'

            # 一级业务编码
            if level1:
                if level1 not in level1_count:
                    level1_count[level1] = len(level1_count) + 1  # 新业务编号递增
                    level2_count = {}
                level1_code = str(level1_count[level1])


            # 二级业务编码
            if level2:
                key = (level1, level2)
                if key not in level2_count:
                    level2_count[key] = len(level2_count) + 1  # 新业务编号递增
                    level3_count = {}
                level2_code = str(level2_count[key])


            # 三级业务编码
            if level3:
                key = (level1, level2, level3)
                if key not in level3_count:
                    level3_count[key] = len(level3_count) + 1  # 新业务编号递增
                    level4_count = {}
                level3_code = str(level3_count[key])


            # 四级业务编码
            if level4:
                key = (level1, level2, level3, level4)
                if key not in level4_count:
                    level4_count[key] = len(level4_count) + 1  # 新业务编号递增
                level4_code = str(level4_count[key])

            # 生成最终编码
            encoding = f"{level1_code}{level2_code}{level3_code}{level4_code}"

            # 将编码写入 AC 列
            sheet[f'{encoding_column}{row[0].row}'] = encoding

        # 保存文件
        wb.save(file_path)
        print("编码生成完成！")

    except PermissionError as e:
        print(f"权限错误：无法保存文件 '{file_path}'。请确保文件未被占用且具有写入权限。")
        print(e)
    except Exception as e:
        print(f"发生意外错误：{e}")

def main():
    """
    主函数：调用生成编码的函数。
    """
    # 指定 Excel 文件路径
    file_path = '项目组_4  开发及其输入和输出--地面-2025.03.18-2（3.0环境导出）--Z 11.13 - 地面总体布局设计 ---最终.xlsx'

    # 调用生成编码的函数
    generate_encoding(file_path)

if __name__ == "__main__":
    main()
