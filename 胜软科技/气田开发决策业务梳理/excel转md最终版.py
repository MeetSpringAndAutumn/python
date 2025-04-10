import shutil
import zipfile
import os
import xml.etree.ElementTree as ET
import openpyxl
import pandas as pd

image_list = []
image_dict = {}


def read_excel_data(filename_path, sheet_name):
    """读取Excel数据并提取图片ID"""
    workbook = openpyxl.load_workbook(filename_path, data_only=False)
    sheet = workbook[sheet_name] if isinstance(
        sheet_name, str) else workbook.worksheets[sheet_name]

    data = []
    for row in sheet.iter_rows(min_row=1, values_only=False):
        row_data = []
        for cell in row:
            if cell.value and isinstance(cell.value, str) and '=_xlfn.DISPIMG(' in cell.value:
                formula = cell.value
                start = formula.find('"') + 1
                end = formula.find('"', start)
                image_id = formula[start:end]
                row_data.append(image_id)
                image_list.append(image_id)
            else:
                row_data.append(cell.value)
        data.append(row_data)
    workbook.close()
    return data


def get_xml_id_image_map(xlsx_file_path):
    """获取XML中图片ID和路径的映射"""
    with zipfile.ZipFile(xlsx_file_path, 'r') as zfile:
        with zfile.open('xl/cellimages.xml') as file:
            xml_content = file.read()
        with zfile.open('xl/_rels/cellimages.xml.rels') as file:
            relxml_content = file.read()

    root = ET.fromstring(xml_content)
    namespaces = {
        'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
        'a': 'http://schemas.openxmlformats.org/drawingml/2006/main'
    }

    name_to_embed_map = {}
    for pic in root.findall('.//xdr:pic', namespaces=namespaces):
        name = pic.find('.//xdr:cNvPr', namespaces=namespaces).attrib['name']
        embed = pic.find('.//a:blip', namespaces=namespaces).attrib[
            '{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed']
        name_to_embed_map[name] = embed

    root1 = ET.fromstring(relxml_content)
    namespaces = {
        'r': 'http://schemas.openxmlformats.org/package/2006/relationships'}

    id_target_map = {child.attrib['Id']: child.attrib.get('Target', 'No Target Found')
                     for child in root1.findall('.//r:Relationship', namespaces=namespaces)}

    name_to_target_map = {name: id_target_map[embed]
                          for name, embed in name_to_embed_map.items()
                          if embed in id_target_map}
    return name_to_target_map


def create_image_dict(xlsx_file_path, sheet_name, output_folder='images'):
    """创建图片ID和文件路径的映射字典"""
    read_excel_data(xlsx_file_path, sheet_name)
    name_to_target_map = get_xml_id_image_map(xlsx_file_path)

    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    with zipfile.ZipFile(xlsx_file_path, 'r') as zfile:
        for image_id in image_list:
            if image_id in name_to_target_map:
                image_path = name_to_target_map[image_id]
                actual_image_path = f'xl/{image_path}'
                if actual_image_path in zfile.namelist():
                    new_file_path = os.path.join(
                        output_folder, f"{image_id}.png")
                    with zfile.open(actual_image_path) as image_file:
                        image_content = image_file.read()
                        with open(new_file_path, 'wb') as new_file:
                            new_file.write(image_content)
                    image_dict[image_id] = f"{output_folder}/{image_id}.png"


def clean_images_folder(folder_path):
    """清空图片文件夹"""
    if os.path.exists(folder_path):
        shutil.rmtree(folder_path)
    os.makedirs(folder_path)


def save_to_md(file_path, content):
    """保存内容到Markdown文件"""
    with open(file_path, 'w', encoding='utf-8') as f:
        f.write(content)


def excel_to_md(excel_file, sheet_name=0, output_folder='images'):
    """将Excel转换为Markdown"""
    global image_list, image_dict
    image_list.clear()
    image_dict.clear()

    create_image_dict(excel_file, sheet_name, output_folder)

    wb = openpyxl.load_workbook(excel_file, data_only=True)
    sheet = wb[sheet_name] if isinstance(
        sheet_name, str) else wb.worksheets[sheet_name]
    merged_ranges = sheet.merged_cells.ranges

    df = pd.read_excel(excel_file, sheet_name=sheet_name, engine='openpyxl')

    column_mapping = {
        '业务场景': '业务场景',
        '一级节点': '一级节点',
        '二级节点': '二级节点',
        '三级节点': '三级节点',
        '末级工作节点': '末级工作节点',
        '产出物名称': '产出物名称',
        '输入物样式': '输入物样式',
        '产出物样式': '产出物样式',
        '处理业务规则': '处理业务规则'
    }

    def get_merged_value(row_idx, col_name):
        try:
            # Convert to 1-based indices
            col_idx = df.columns.get_loc(col_name) + 1
            row_idx = row_idx + 2  # Excel row is 1-based and has header

            for merged_range in merged_ranges:
                if (merged_range.min_row <= row_idx <= merged_range.max_row and
                        merged_range.min_col <= col_idx <= merged_range.max_col):
                    cell_value = sheet.cell(
                        merged_range.min_row, merged_range.min_col).value
                    return str(cell_value) if cell_value is not None else ''
            return None
        except Exception as e:
            print(
                f"Error in get_merged_value: row={row_idx}, col={col_name}, error={e}")
            return None

    # 预处理节点产出物计数
    node_output_counts = {}  # 存储每个节点路径的产出物数量
    node_outputs = {}  # 存储每个节点路径的产出物列表

    for i, r in df.iterrows():
        # 获取各级节点值
        node_values = {}
        for level in ['业务场景', '一级节点', '二级节点', '三级节点', '末级工作节点']:
            merged_value = get_merged_value(i, column_mapping[level])
            if merged_value is not None:
                node_values[level] = merged_value
            else:
                value = r[column_mapping[level]]
                node_values[level] = str(value) if not pd.isna(value) else ''

        # 构建节点的完整路径，包含所有非空节点
        node_path = []
        for level in ['业务场景', '一级节点', '二级节点', '三级节点', '末级工作节点']:
            if node_values[level]:
                node_path.append((level, node_values[level]))

        # 将列表转换为元组以便可以作为字典键
        node_path_tuple = tuple(node_path)

        # 找到最后一个非空节点
        last_level = None
        last_content = None
        for level in ['末级工作节点', '三级节点', '二级节点', '一级节点', '业务场景']:
            if node_values[level]:
                last_level = level
                last_content = node_values[level]
                break

        # 获取产出物
        output_name = str(r[column_mapping['产出物名称']]) if not pd.isna(
            r[column_mapping['产出物名称']]) else ''

        # 如果有产出物和有效的节点路径
        if output_name and node_path_tuple:
            # 初始化节点路径的产出物列表
            if node_path_tuple not in node_outputs:
                node_outputs[node_path_tuple] = []
                node_output_counts[node_path_tuple] = 0

            # 如果这个产出物还未记录，添加到列表并增加计数
            if output_name not in node_outputs[node_path_tuple]:
                node_outputs[node_path_tuple].append(output_name)
                node_output_counts[node_path_tuple] += 1

    markdown_content = ""
    last_output = {key: '' for key in [
        '业务场景', '一级节点', '二级节点', '三级节点', '末级工作节点']}

    for index, row in df.iterrows():
        current_nodes = {}

        # 获取当前行的所有节点值
        for node_type in ['业务场景', '一级节点', '二级节点', '三级节点', '末级工作节点']:
            merged_value = get_merged_value(index, column_mapping[node_type])
            if merged_value is not None:
                current_nodes[node_type] = merged_value
            else:
                value = row[column_mapping[node_type]]
                current_nodes[node_type] = str(
                    value) if not pd.isna(value) else ''

        # 构建当前行节点的完整路径
        current_path = []
        for level in ['业务场景', '一级节点', '二级节点', '三级节点', '末级工作节点']:
            if current_nodes[level]:
                current_path.append((level, current_nodes[level]))
        current_path_tuple = tuple(current_path)

        # 改进的节点去重逻辑：检查所有节点的内容重复
        node_values = set()  # 用于存储节点内容，检测重复
        # 先从高级别节点开始处理
        for level in ['业务场景', '一级节点', '二级节点', '三级节点', '末级工作节点']:
            if current_nodes[level] and current_nodes[level] in node_values:
                # 如果节点内容已经在更高级别出现过，则清空当前节点
                current_nodes[level] = ''
            elif current_nodes[level]:
                # 否则，将当前节点内容添加到已处理集合中
                node_values.add(current_nodes[level])

        # 检查当前行是否需要输出各级标题
        for level in ['业务场景', '一级节点', '二级节点', '三级节点', '末级工作节点']:
            # 如果当前节点为空，跳过
            if not current_nodes[level]:
                continue

            # 判断是否需要输出该节点
            output_this_node = False

            # 如果节点内容与上次输出的不同，需要输出
            if current_nodes[level] != last_output[level]:
                output_this_node = True

            # 检查上一层级节点是否发生变化，如果是则当前节点也需要输出
            if level == '一级节点' and current_nodes['业务场景'] != last_output['业务场景']:
                output_this_node = True
            elif level == '二级节点' and (current_nodes['一级节点'] != last_output['一级节点']):
                output_this_node = True
            elif level == '三级节点' and (current_nodes['二级节点'] != last_output['二级节点']):
                output_this_node = True
            elif level == '末级工作节点' and (current_nodes['三级节点'] != last_output['三级节点']):
                output_this_node = True

            # 如果需要输出该节点
            if output_this_node:
                # 根据层级输出标题
                if level == '业务场景':
                    markdown_content += f"# {current_nodes[level]}\n"
                elif level == '一级节点':
                    markdown_content += f"## {current_nodes[level]}\n"
                elif level == '二级节点':
                    markdown_content += f"### {current_nodes[level]}\n"
                elif level == '三级节点':
                    markdown_content += f"#### {current_nodes[level]}\n"
                elif level == '末级工作节点':
                    markdown_content += f"##### {current_nodes[level]}\n"

                # 更新已输出的节点记录
                last_output[level] = current_nodes[level]

                # 如果上层节点变化，需要重置所有下层节点的输出记录
                if level == '业务场景':
                    last_output['一级节点'] = ''
                    last_output['二级节点'] = ''
                    last_output['三级节点'] = ''
                    last_output['末级工作节点'] = ''
                elif level == '一级节点':
                    last_output['二级节点'] = ''
                    last_output['三级节点'] = ''
                    last_output['末级工作节点'] = ''
                elif level == '二级节点':
                    last_output['三级节点'] = ''
                    last_output['末级工作节点'] = ''
                elif level == '三级节点':
                    last_output['末级工作节点'] = ''

        # 处理产出物
        output_name = str(row[column_mapping['产出物名称']]) if not pd.isna(
            row[column_mapping['产出物名称']]) else ''

        # 找到当前行的最后一个非空节点
        last_level = None
        last_content = None
        for level in ['末级工作节点', '三级节点', '二级节点', '一级节点', '业务场景']:
            if current_nodes[level]:
                last_level = level
                last_content = current_nodes[level]
                break

        # 检查当前节点是否有多个不同的产出物
        has_multiple_outputs = current_path_tuple in node_output_counts and node_output_counts[
            current_path_tuple] > 1

        # 决定是否输出产出物名称
        if output_name:
            # 如果产出物名称与节点不同，或者该节点有多个产出物，则输出产出物
            if output_name != last_content or has_multiple_outputs:
                markdown_content += f"- {output_name}\n"

        # 处理输入物样式
        input_style = str(row[column_mapping['输入物样式']]) if not pd.isna(
            row[column_mapping['输入物样式']]) else ''
        if input_style:
            markdown_content += "  - 输入物\n"
            if '=DISPIMG' in input_style:
                img_id = input_style[input_style.find(
                    '"') + 1:input_style.rfind('"')]
                if img_id in image_dict:
                    markdown_content += f"    - ![](images/{img_id}.png)\n"
                else:
                    markdown_content += f"    - {input_style}\n"
            else:
                markdown_content += f"    - {input_style}\n"

        # 处理产出物样式
        output_style = str(row[column_mapping['产出物样式']]) if not pd.isna(
            row[column_mapping['产出物样式']]) else ''
        if output_style:
            markdown_content += "  - 产出物样式\n"
            if '=DISPIMG' in output_style:
                img_id = output_style[output_style.find(
                    '"') + 1:output_style.rfind('"')]
                if img_id in image_dict:
                    markdown_content += f"    - ![](images/{img_id}.png)\n"
                else:
                    markdown_content += f"    - {output_style}\n"
            else:
                markdown_content += f"    - {output_style}\n"

        # 处理业务规则
        processing_rules = str(row[column_mapping['处理业务规则']]) if not pd.isna(
            row[column_mapping['处理业务规则']]) else ''
        if processing_rules and processing_rules != 'nan':
            markdown_content += f"  - 处理业务规则\n    - {processing_rules}\n"

    wb.close()
    return markdown_content


if __name__ == '__main__':
    excel_file = '安全环保IPOM表-20250224-生董董.xlsx'
    base_output_folder = 'output'
    images_folder = os.path.join(base_output_folder, 'images')

    try:
        if not os.path.exists(base_output_folder):
            os.makedirs(base_output_folder)

        clean_images_folder(images_folder)

        wb = openpyxl.load_workbook(excel_file, data_only=False)
        sheet_names = wb.sheetnames
        wb.close()

        for sheet_name in sheet_names:
            markdown_file = os.path.join(
                base_output_folder, f'{sheet_name}.md')
            print(f"正在处理工作表: {sheet_name}")
            md_content = excel_to_md(
                excel_file, sheet_name=sheet_name, output_folder=images_folder)
            save_to_md(markdown_file, md_content)
            print(f"已保存工作表 {sheet_name} 的内容到 {markdown_file}")

        print("所有工作表处理完成")
        print(f"输出目录: {os.path.abspath(base_output_folder)}")

    except Exception as e:
        print(f"发生错误：{e}")
        print("请检查Excel文件中的列名是否正确")
